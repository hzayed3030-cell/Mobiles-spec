"""
export_service.py
خدمات تصدير البيانات إلى Excel و PDF وواجهات الطباعة Print مع دعم RTL الكامل
وإبراز خلية الهاتف الفائز باللون الأخضر مع مربع "⭐ الأفضل" وعمود "سبب الأفضلية".
"""

import io
import os
import datetime
from typing import List, Dict, Any, Optional
import pandas as pd

# OpenPyXL لتصدير ملفات Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab لتصدير ملفات PDF
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# تشكيل النصوص العربية و bidi
import arabic_reshaper
from bidi.algorithm import get_display


# --- تسجيل الخطوط العربية في ReportLab ---
ARABIC_FONT_NAME = "ArabicFont"
ARABIC_FONT_BOLD = "ArabicFontBold"

def register_arabic_fonts():
    font_candidates = [
        ("C:/Windows/Fonts/tahoma.ttf", "C:/Windows/Fonts/tahomabd.ttf"),
        ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
        ("C:/Windows/Fonts/segoeui.ttf", "C:/Windows/Fonts/segoeuib.ttf")
    ]
    registered = False
    for regular, bold in font_candidates:
        if os.path.exists(regular):
            try:
                pdfmetrics.registerFont(TTFont(ARABIC_FONT_NAME, regular))
                if os.path.exists(bold):
                    pdfmetrics.registerFont(TTFont(ARABIC_FONT_BOLD, bold))
                else:
                    pdfmetrics.registerFont(TTFont(ARABIC_FONT_BOLD, regular))
                registered = True
                break
            except Exception:
                continue
    if not registered:
        pdfmetrics.registerFont(TTFont(ARABIC_FONT_NAME, "Helvetica"))
        pdfmetrics.registerFont(ARABIC_FONT_BOLD, "Helvetica-Bold")

register_arabic_fonts()


def fix_arabic_text(text: Any) -> str:
    """معالجة النص العربي بالتشكيل وتعديل الاتجاه Bidi"""
    if text is None:
        return ""
    text_str = str(text)
    if not text_str.strip():
        return ""
    try:
        has_arabic = any('\u0600' <= char <= '\u06FF' or '\u0750' <= char <= '\u077F' for char in text_str)
        if has_arabic:
            reshaped_text = arabic_reshaper.reshape(text_str)
            bidi_text = get_display(reshaped_text)
            return bidi_text
        return text_str
    except Exception:
        return text_str


# ==============================================================================
# 1. EXCEL EXPORT (RTL with Green Winner Cells + [الأفضل] Box + Reason Column)
# ==============================================================================

def export_comparison_to_excel(
    comparison_df: pd.DataFrame,
    title: str = "مقارنة مواصفات الموبايلات",
    winners_per_row: Optional[List[str]] = None,
    final_verdict: Optional[Dict[str, Any]] = None
) -> bytes:
    """
    تصدير جدول المقارنة إلى ملف Excel مع تلوين الخلية الفائزة بالأخضر وإضافة شارة الأفضل وعمود سبب الأفضلية.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "مقارنة المواصفات"

    # تفعيل اتجاه الورقة وتخطيط الصفحة من اليسار إلى اليمين (Page Layout -> Sheet Left-to-Right)
    ws.views.sheetView[0].rightToLeft = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # إعداد الأنماط والألوان
    font_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="CBD5E1")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_spec_name = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    font_cell = Font(name="Segoe UI", size=10, color="334155")
    font_winner_cell = Font(name="Segoe UI", size=10, bold=True, color="14532D")
    font_reason_cell = Font(name="Segoe UI", size=10, bold=True, color="1E40AF")
    font_verdict_label = Font(name="Segoe UI", size=11, bold=True, color="78350F")

    fill_title = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_header_reason = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_spec_col = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_winner_cell = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_reason_col = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    fill_row_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_row_odd = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_verdict = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    thick_verdict_border = Border(
        left=Side(style='medium', color='F59E0B'),
        right=Side(style='medium', color='F59E0B'),
        top=Side(style='medium', color='F59E0B'),
        bottom=Side(style='medium', color='F59E0B')
    )

    num_cols = len(comparison_df.columns)

    # 1. عنوان التقرير
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell_title = ws.cell(row=1, column=1, value=title)
    cell_title.font = font_title
    cell_title.fill = fill_title
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 2. تاريخ التقرير
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    cell_sub = ws.cell(row=2, column=1, value=f"تاريخ التصدير: {now_str} | الخلايا الخضراء تمثل القيمة المتفوقة [الأفضل] | عمود سبب الأفضلية يوضح سر التفوق")
    cell_sub.font = font_subtitle
    cell_sub.fill = fill_title
    cell_sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 10

    # 3. رؤوس الأعمدة
    header_row_idx = 4
    ws.row_dimensions[header_row_idx].height = 30

    for col_idx, col_name in enumerate(comparison_df.columns, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_header_reason if col_idx == num_cols else fill_header
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 4. صفوف المواصفات
    last_data_row = 4
    for row_idx, (_, row_data) in enumerate(comparison_df.iterrows(), start=5):
        ws.row_dimensions[row_idx].height = 32
        last_data_row = row_idx
        is_odd = (row_idx % 2 == 1)
        row_fill = fill_row_odd if is_odd else fill_row_even

        spec_row_idx = row_idx - 5
        winner_phone = winners_per_row[spec_row_idx] if (winners_per_row and spec_row_idx < len(winners_per_row)) else None

        for col_idx, val in enumerate(row_data, start=1):
            col_name = comparison_df.columns[col_idx - 1]
            cell_val_str = str(val)

            if col_idx == 1:
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_val_str)
                cell.font = font_spec_name
                cell.fill = fill_spec_col
            elif col_idx == num_cols and col_name == "سبب الأفضلية":
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_val_str)
                cell.font = font_reason_cell
                cell.fill = fill_reason_col
            else:
                if winner_phone and col_name == winner_phone and winner_phone != "-":
                    cell = ws.cell(row=row_idx, column=col_idx, value=f"{cell_val_str}\n[⭐ الأفضل]")
                    cell.font = font_winner_cell
                    cell.fill = fill_winner_cell
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_val_str)
                    cell.font = font_cell
                    cell.fill = row_fill

            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if col_idx in [1, num_cols] else "center", vertical="center", wrap_text=True)

    # 5. صف التقييم النهائي
    if final_verdict and final_verdict.get("winner_name"):
        verdict_row_idx = last_data_row + 1
        ws.row_dimensions[verdict_row_idx].height = 36

        c_label = ws.cell(row=verdict_row_idx, column=1, value="🏆 التقييم النهائي")
        c_label.font = font_verdict_label
        c_label.fill = fill_verdict
        c_label.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        c_label.border = thick_verdict_border

        winner_name = final_verdict.get("winner_name", "")
        for col_idx in range(2, num_cols + 1):
            col_name = comparison_df.columns[col_idx - 1]
            cell = ws.cell(row=verdict_row_idx, column=col_idx)
            cell.fill = fill_verdict
            cell.border = thick_verdict_border

            if col_name == "سبب الأفضلية":
                cell.value = "الأكثر تفوقاً في المواصفات"
                cell.font = font_reason_cell
            elif col_name == winner_name:
                cell.value = "👑 الخيار الأفضل إجمالاً"
                cell.font = Font(name="Segoe UI", size=11, bold=True, color="14532D")
            else:
                score = final_verdict.get("all_scores", {}).get(col_name, 0)
                cell.value = f"فاز في ({score}) مواصفات"
                cell.font = font_cell
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 6. ضبط عرض الأعمدة
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            ws.column_dimensions[col_letter].width = 25
        elif col_idx == num_cols and comparison_df.columns[col_idx - 1] == "سبب الأفضلية":
            ws.column_dimensions[col_letter].width = 30
        else:
            max_len = max(len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(4, ws.max_row + 1))
            col_width = min(max(max_len + 4, 25), 45)
            ws.column_dimensions[col_letter].width = col_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==============================================================================
# 2. PDF EXPORT (RTL Supported with Green Winner Cells + [الأفضل] Box)
# ==============================================================================

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super(NumberedCanvas, self).__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super(NumberedCanvas, self).showPage()
        super(NumberedCanvas, self).save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont(ARABIC_FONT_NAME, 9)
        self.setFillColor(colors.HexColor("#64748B"))
        footer_text = fix_arabic_text(f"مقارنة مواصفات الهواتف | صفحة {self._pageNumber} من {page_count}")
        self.drawRightString(A4[1] - 40 if self._pagesize == landscape(A4) else A4[0] - 40, 20, footer_text)
        self.restoreState()


def export_comparison_to_pdf(
    comparison_df: pd.DataFrame,
    title: str = "مقارنة مواصفات الهواتف الذكية",
    winners_per_row: Optional[List[str]] = None,
    final_verdict: Optional[Dict[str, Any]] = None
) -> bytes:
    """
    تصدير جدول المقارنة إلى ملف PDF منظم مع تلوين خلايا القيم الفائزة بالأخضر وعمود سبب الأفضلية.
    """
    buffer = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=30
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        name="ArabicTitle", fontName=ARABIC_FONT_BOLD, fontSize=16, leading=22, alignment=1, textColor=colors.HexColor("#1E3A8A"), spaceAfter=4
    )
    sub_style = ParagraphStyle(
        name="ArabicSubTitle", fontName=ARABIC_FONT_NAME, fontSize=9, leading=12, alignment=1, textColor=colors.HexColor("#64748B"), spaceAfter=14
    )
    cell_style = ParagraphStyle(
        name="TableCell", fontName=ARABIC_FONT_NAME, fontSize=8.5, leading=11, alignment=1, textColor=colors.HexColor("#1E293B")
    )
    cell_winner_style = ParagraphStyle(
        name="TableWinnerCell", fontName=ARABIC_FONT_BOLD, fontSize=8.5, leading=11, alignment=1, textColor=colors.HexColor("#14532D")
    )
    cell_reason_style = ParagraphStyle(
        name="TableReasonCell", fontName=ARABIC_FONT_BOLD, fontSize=8.5, leading=11, alignment=1, textColor=colors.HexColor("#1E40AF")
    )
    cell_spec_style = ParagraphStyle(
        name="TableSpecCell", fontName=ARABIC_FONT_BOLD, fontSize=8.5, leading=11, alignment=2, textColor=colors.HexColor("#0F172A")
    )
    header_style = ParagraphStyle(
        name="TableHeader", fontName=ARABIC_FONT_BOLD, fontSize=9.5, leading=12, alignment=1, textColor=colors.white
    )
    verdict_style = ParagraphStyle(
        name="TableVerdictCell", fontName=ARABIC_FONT_BOLD, fontSize=9, leading=12, alignment=1, textColor=colors.HexColor("#78350F")
    )

    story.append(Paragraph(fix_arabic_text(title), title_style))
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    story.append(Paragraph(fix_arabic_text(f"تاريخ التقرير: {now_str} | الخلايا الخضراء تمثل القيمة المتفوقة [⭐ الأفضل] | عمود سبب الأفضلية يوضح سر التفوق"), sub_style))

    table_data = []
    header_row = [Paragraph(fix_arabic_text(col), header_style) for col in comparison_df.columns]
    table_data.append(header_row)

    table_styles_list = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor("#F1F5F9")),
    ]

    num_cols = len(comparison_df.columns)

    for row_idx, (_, row) in enumerate(comparison_df.iterrows(), start=1):
        spec_row_idx = row_idx - 1
        winner_phone = winners_per_row[spec_row_idx] if (winners_per_row and spec_row_idx < len(winners_per_row)) else None
        row_cells = []

        for col_idx, val in enumerate(row):
            col_name = comparison_df.columns[col_idx]
            formatted_text = fix_arabic_text(str(val))
            
            if col_idx == 0:
                row_cells.append(Paragraph(formatted_text, cell_spec_style))
            elif col_idx == num_cols - 1 and col_name == "سبب الأفضلية":
                row_cells.append(Paragraph(formatted_text, cell_reason_style))
                table_styles_list.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor("#EFF6FF")))
            else:
                if winner_phone and col_name == winner_phone and winner_phone != "-":
                    best_badge_str = fix_arabic_text("[⭐ الأفضل]")
                    combined_text = f"{formatted_text}<br/><font color='#15803D'><b>{best_badge_str}</b></font>"
                    row_cells.append(Paragraph(combined_text, cell_winner_style))
                    table_styles_list.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor("#DCFCE7")))
                else:
                    row_cells.append(Paragraph(formatted_text, cell_style))

        table_data.append(row_cells)

    # صف التقييم النهائي
    if final_verdict and final_verdict.get("winner_name"):
        winner_name = final_verdict.get("winner_name", "")
        v_cells = []
        for col_idx in range(len(comparison_df.columns)):
            col_name = comparison_df.columns[col_idx]
            if col_idx == 0:
                v_cells.append(Paragraph(fix_arabic_text("التقييم النهائي"), verdict_style))
            elif col_name == "سبب الأفضلية":
                v_cells.append(Paragraph(fix_arabic_text("الأكثر تفوقاً"), cell_reason_style))
            else:
                if col_name == winner_name:
                    v_cells.append(Paragraph(fix_arabic_text("الخيار الأفضل إجمالاً 👑"), verdict_style))
                else:
                    score = final_verdict.get("all_scores", {}).get(col_name, 0)
                    v_cells.append(Paragraph(fix_arabic_text(f"فاز بـ ({score}) مواصفة"), cell_style))
        table_data.append(v_cells)
        table_styles_list.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#FEF3C7")))
        table_styles_list.append(('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.HexColor("#F59E0B")))

    page_width = page_size[0] - 40
    first_col_w = max(100, int(page_width * 0.20))
    last_col_w = max(110, int(page_width * 0.22)) if comparison_df.columns[-1] == "سبب الأفضلية" else 0
    phone_cols_count = num_cols - (2 if last_col_w else 1)
    other_col_w = int((page_width - first_col_w - last_col_w) / phone_cols_count) if phone_cols_count > 0 else 100
    
    col_widths = [first_col_w] + [other_col_w] * phone_cols_count
    if last_col_w:
        col_widths.append(last_col_w)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(table_styles_list))
    story.append(table)

    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer.getvalue()


# ==============================================================================
# 3. HTML PRINT VIEW (Print / Web Preview)
# ==============================================================================

def generate_html_print_view(
    comparison_df: pd.DataFrame,
    title: str = "مقارنة مواصفات الهواتف الذكية",
    winners_per_row: Optional[List[str]] = None,
    final_verdict: Optional[Dict[str, Any]] = None,
    reasons_per_row: Optional[List[str]] = None,
    **kwargs
) -> str:
    """
    توليد صفحة HTML احترافية للطباعة مع تلوين خلايا القيم الفائزة بالأخضر مع مربع "⭐ الأفضل" وعمود سبب الأفضلية.
    """
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    headers_html = "".join([f"<th>{col}</th>" for col in comparison_df.columns])
    num_cols = len(comparison_df.columns)

    rows_html = []
    for r_idx, (_, row) in enumerate(comparison_df.iterrows()):
        winner_phone = winners_per_row[r_idx] if (winners_per_row and r_idx < len(winners_per_row)) else None
        row_str = "<tr>"
        for i, val in enumerate(row):
            col_name = comparison_df.columns[i]
            if i == 0:
                row_str += f"<td class='spec-name'>{val}</td>"
            elif i == num_cols - 1 and col_name == "سبب الأفضلية":
                row_str += f"<td class='reason-cell'>{val}</td>"
            else:
                if winner_phone and col_name == winner_phone and winner_phone != "-":
                    row_str += f"<td class='winner-cell'>{val}<div class='best-box-badge'>⭐ الأفضل</div></td>"
                else:
                    row_str += f"<td>{val}</td>"
        row_str += "</tr>"
        rows_html.append(row_str)

    if final_verdict and final_verdict.get("winner_name"):
        winner_name = final_verdict.get("winner_name", "")
        v_row = "<tr class='verdict-row'>"
        for i in range(len(comparison_df.columns)):
            col_name = comparison_df.columns[i]
            if i == 0:
                v_row += f"<td class='spec-name'>🏆 التقييم النهائي</td>"
            elif col_name == "سبب الأفضلية":
                v_row += f"<td class='reason-cell'><strong>الأكثر تفوقاً</strong></td>"
            else:
                if col_name == winner_name:
                    v_row += f"<td class='winner-cell'><strong>👑 الخيار الأفضل إجمالاً</strong></td>"
                else:
                    score = final_verdict.get("all_scores", {}).get(col_name, 0)
                    v_row += f"<td>فاز في ({score}) مواصفات</td>"
        v_row += "</tr>"
        rows_html.append(v_row)

    tbody_html = "\n".join(rows_html)

    html_content = f"""
    <!DOCTYPE html>
    <html lang="ar" dir="rtl">
    <head>
        <meta charset="UTF-8">
        <title>{title}</title>
        <link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700;800;900&display=swap" rel="stylesheet">
        <style>
            body {{
                font-family: 'Cairo', sans-serif;
                background-color: #f8fafc;
                color: #0f172a;
                direction: rtl;
                text-align: right;
                padding: 20px;
            }}
            .print-container {{
                max-width: 1250px;
                margin: 0 auto;
                background: #ffffff;
                border-radius: 14px;
                padding: 24px;
                box-shadow: 0 4px 16px rgba(0,0,0,0.06);
            }}
            .action-bar {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin-bottom: 20px;
                border-bottom: 2px solid #e2e8f0;
                padding-bottom: 12px;
            }}
            .btn-print {{
                background: #1e40af;
                color: #ffffff;
                border: none;
                padding: 10px 24px;
                font-size: 15px;
                font-weight: 700;
                font-family: 'Cairo', sans-serif;
                border-radius: 8px;
                cursor: pointer;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                text-align: center;
                font-size: 13px;
            }}
            th {{
                background-color: #1e3a8a;
                color: #ffffff;
                padding: 12px;
                border: 1px solid #1e3a8a;
            }}
            th:first-child {{
                background-color: #0f172a;
                text-align: right;
                width: 20%;
            }}
            th:last-child {{
                background-color: #0f172a;
                width: 22%;
            }}
            td {{
                padding: 10px 12px;
                border: 1px solid #cbd5e1;
                vertical-align: middle;
            }}
            td.spec-name {{
                font-weight: 700;
                background-color: #f1f5f9;
                text-align: right;
            }}
            td.winner-cell {{
                background-color: #dcfce7 !important;
                color: #14532d !important;
                font-weight: 800 !important;
                border: 2px solid #10b981 !important;
            }}
            .best-box-badge {{
                background: #15803d;
                color: #ffffff;
                padding: 2px 8px;
                border-radius: 6px;
                font-size: 11px;
                font-weight: 800;
                display: inline-block;
                margin-top: 4px;
            }}
            td.reason-cell {{
                background-color: #eff6ff !important;
                color: #1e40af !important;
                font-weight: 700 !important;
            }}
            .verdict-row {{
                background-color: #fef3c7 !important;
            }}
            @media print {{
                body {{ padding: 0 !important; background: #fff !important; }}
                .action-bar {{ display: none !important; }}
                @page {{ size: A4 landscape; margin: 10mm; }}
            }}
        </style>
    </head>
    <body>
        <div class="print-container">
            <div class="action-bar">
                <strong>أداة مقارنة الهواتف الذكية</strong>
                <button class="btn-print" onclick="window.print()">🖨️ طباعة (Print)</button>
            </div>
            <h2 style="text-align:center; color:#1e3a8a;">{title}</h2>
            <p style="text-align:center; color:#64748b; font-size:12px; margin-bottom:16px;">تاريخ التقرير: {now_str} | الخلايا الخضراء تمثل القيمة المتفوقة [⭐ الأفضل] | عمود سبب الأفضلية يوضح سر التفوق</p>
            <table>
                <thead><tr>{headers_html}</tr></thead>
                <tbody>{tbody_html}</tbody>
            </table>
        </div>
    </body>
    </html>
    """
    return html_content


# ==============================================================================
# 4. UNIVERSAL PHONE SETTINGS EXCEL EXPORT (Left-to-Right Layout)
# ==============================================================================

# كتالوج إعدادات آبل (Apple iOS 18/19 Settings Catalog)
APPLE_IOS_SETTINGS = [
    [1, "Apple Intelligence & Siri", "Apple Intelligence", "Writing Tools & Clean Up", "AI-powered text rewriting, tone adjusting, summarization, and photo object cleanup", "إعادة صياغة وتلخيص النصوص بالذكاء الاصطناعي وإزالة العناصر من الصور", "Settings > Apple Intelligence & Siri > Writing Tools"],
    [1, "Apple Intelligence & Siri", "Siri", "Siri Requests & Language", "Configure Siri voice, language preferences, and Type to Siri interface", "ضبط لغة وتفضيلات المساعد الصوتي وتفعيل نمط الكتابة لـ Siri", "Settings > Apple Intelligence & Siri > Siri Requests"],
    [2, "Wi-Fi", "Wi-Fi Networks", "Intelligent Auto-Join", "Auto-connect to Wi-Fi networks and manage Private Wi-Fi MAC addressing", "الاتصال التلقائي بشبكات الواي فاي وإدارة العناوين الخاصة Private Wi-Fi", "Settings > Wi-Fi"],
    [2, "Wi-Fi", "Wi-Fi Calling", "Cellular Preferred over Wi-Fi", "Make and receive crystal-clear cellular voice calls over Wi-Fi connection", "إجراء المكالمات الخلوية عبر شبكة الواي فاي عند ضعف التغطية", "Settings > Cellular > Wi-Fi Calling"],
    [3, "Cellular", "Cellular Data Options", "Voice & Data (5G Auto / 5G On)", "Toggle 5G Standalone mode, data roaming, and multi-eSIM carrier management", "تحديد نمط شبكة الجيل الخامس وإدارة تجوال البيانات وخطوط eSIM", "Settings > Cellular > Cellular Data Options"],
    [3, "Cellular", "Personal Hotspot", "Allow Others to Join", "Share high-speed mobile cellular data as a password-protected Wi-Fi hotspot", "مشاركة باقة الإنترنت كنقطة اتصال واي فاي محمية بكلمة مرور", "Settings > Cellular > Personal Hotspot"],
    [4, "Bluetooth", "My Devices", "Device Disconnect & Audio Sharing", "Manage paired Bluetooth accessories and share audio with another AirPods pair", "إدارة ملحقات البلوتوث ومشاركة الصوت مع سماعات AirPods أخرى", "Settings > Bluetooth"],
    [5, "Notifications", "Scheduled Summary", "Delivery Times", "Bundle non-urgent notifications into convenient morning and evening digests", "جدولة وصول الإشعارات غير العاجلة في ملخصات صباحية ومسائية مجمعة", "Settings > Notifications > Scheduled Summary"],
    [6, "Sounds & Haptics", "Haptic Feedback", "Keyboard Haptics / System Haptics", "Enable tactile vibrational feedback for virtual keyboard typing and system gestures", "تفعيل الاهتزازات التفاعلية الدقيقة للوحة المفاتيح والنظام", "Settings > Sounds & Haptics"],
    [6, "Sounds & Haptics", "Headphone Safety", "Reduce Loud Sounds", "Monitor ear health exposure and cap maximum audio decibel output limit", "حماية حاسة السمع وتحديد الحد الأقصى لمستوى الديسيبل المسموح", "Settings > Sounds & Haptics > Headphone Safety"],
    [7, "Focus", "Focus Modes", "Do Not Disturb / Work / Sleep", "Create customized Focus profiles to filter notifications and match Lock Screens", "إنشاء بروفايلات تركيز مخصصة لحجب الإشعارات وضبط شاشات القفل", "Settings > Focus"],
    [8, "Screen Time", "Downtime & App Limits", "Category Limits", "Schedule time away from screen and set daily usage limits for app categories", "تحديد أوقات الابتعاد عن الشاشة ووضع حدود زمنية لاستخدام التطبيقات", "Settings > Screen Time > App Limits"],
    [9, "General", "Software Update", "Automatic Updates (iOS / Security)", "Download and install iOS system updates and Rapid Security Response patches", "تنزيل وتثبيت تحديثات نظام iOS وتحديثات الاستجابة الأمنية السريعة", "Settings > General > Software Update"],
    [9, "General", "AirDrop", "NameDrop / Everyone for 10 Minutes", "Transfer files and share contact info instantly by bringing iPhones close together", "تبادل الملفات وجهات الاتصال بمجرد تقريب هواتف الآيفون من بعضها", "Settings > General > AirDrop"],
    [9, "General", "AirPlay & Handoff", "Continuity Camera & Handoff", "Seamlessly resume work across Apple devices and use iPhone as a Mac webcam", "استكمال تصفح الويب والعمل واستخدام الآيفون ككاميرا ويب للماك", "Settings > General > AirPlay & Handoff"],
    [9, "General", "iPhone Storage", "Offload Unused Apps", "Free up device storage by removing unused apps while preserving user data", "إدارة الذاكرة وحذف التطبيقات غير المستخدمة مع الاحتفاظ ببياناتها", "Settings > General > iPhone Storage"],
    [10, "Display & Brightness", "Appearance", "Light / Dark / Auto", "Switch automatically between Light and Dark mode based on sunset schedule", "التبديل بين النمط الفاتح والداكن تلقائياً مع غروب وشروق الشمس", "Settings > Display & Brightness > Appearance"],
    [10, "Display & Brightness", "Always On Display", "Show Wallpaper / Show Notifications", "Customize Always On lock screen behavior and adaptive ProMotion 120Hz refresh", "تخصيص شاشة التشغيل الدائم وتقنية ProMotion 120Hz التكيفية", "Settings > Display & Brightness > Always On Display"],
    [11, "Action Button", "Action Shortcut", "Camera / Focus / Shortcut / Voice Memo", "Map side Action button to launch favorite tools, shortcuts, or camera modes instantly", "تخصيص زر الإجراءات الجانبي لتنفيذ مهام فورية بضغطة واحدة", "Settings > Action Button"],
    [12, "StandBy", "Display Mode", "Night Mode (Red Tint)", "Transform charging iPhone into a smart bedside clock and customizable widget display", "عرض الساعة والتقويم والويدجت بذكاء أثناء شحن الهاتف أفقياً", "Settings > StandBy"],
    [13, "Face ID & Passcode", "Biometrics", "Face ID with a Mask / Require Attention", "3D biometric authentication with eye-attention sensing and mask support", "فتح القفل ببصمة الوجه المتقدمة ثلاثية الأبعاد والتأكد من انتباه العين", "Settings > Face ID & Passcode"],
    [13, "Face ID & Passcode", "Stolen Device Protection", "Security Delay", "Enforce biometric security and delay passcode resets when away from trusted locations", "طلب المصادقة الحيوية ومنع تغيير كلمة السر عند الابتعاد عن المواقع المألوفة", "Settings > Face ID & Passcode > Stolen Device Protection"],
    [14, "Battery", "Battery Health & Charging", "Charging Optimization (80% Limit)", "View maximum battery capacity and cap charging at 80% to prolong chemical health", "معاينة السعة القصوى للبطارية وتحديد سقف الشحن عند 80% للحفاظ عليها", "Settings > Battery > Battery Health & Charging"],
    [15, "Privacy & Security", "Safety Check", "Emergency Reset & Review Sharing", "Instantly revoke permissions and review people/apps with access to your private info", "المراجعة الفورية للصلاحيات والأشخاص والتطبيقات التي تشارك معهم بياناتك", "Settings > Privacy & Security > Safety Check"],
    [15, "Privacy & Security", "Lockdown Mode", "Extreme Protection", "Extreme security mode protecting against sophisticated targeted cyber attacks", "وضع الحماية القصوى لحجب الهجمات السيبرانية وبرمجيات التجسس المتقدمة", "Settings > Privacy & Security > Lockdown Mode"],
    [16, "Camera", "Formats", "ProRAW & ProRes Video", "Capture 48MP ProRAW photos and cinematic professional ProRes 4K Log videos", "التقاط صور بدقة 48MP ProRAW وتصوير فيديو سينمائي احترافي 4K Log", "Settings > Camera > Formats"],
    [16, "Camera", "Photographic Styles", "Tone & Warmth Customization", "Personalize camera rendering pipeline with custom color tones and skin warmths", "تخصيص نمط ألوان الصور الافتراضي ومعالجة درجات لون البشرة والظلال", "Settings > Camera > Photographic Styles"],
    [17, "Accessibility", "Vision & Hearing", "VoiceOver / Live Captions / AssistiveTouch", "Universal accessibility features including real-time audio captioning and touch assistance", "تسهيلات الاستخدام وتحويل الكلام لنصوص مكتوبة لحظياً للمكالمات والوسائط", "Settings > Accessibility"],
    [18, "Emergency SOS", "Crash Detection", "Call with 5 Presses / Hold & Release", "Automatic emergency services dialing and GPS beacon upon detecting severe vehicle crash", "الاتصال التلقائي بالطوارئ وإرسال الموقع عند استشعار حوادث السيارات", "Settings > Emergency SOS"],
]

# كتالوج إعدادات شاومي وهواوي (Xiaomi HyperOS / EMUI Settings Catalog)
XIAOMI_HYPEROS_SETTINGS = [
    [1, "SIM cards & mobile networks", "5G Network", "Smart 5G / VoLTE", "Enable smart 5G auto-switch, data roaming, and dual-SIM management", "تفعيل شبكات الجيل الخامس الذكية وإدارة تجوال البيانات والشرائح", "Settings > SIM cards & mobile networks"],
    [2, "Wi-Fi & Internet", "Wi-Fi Assistant", "Dual-band Wi-Fi speed boost", "Combine 2.4GHz and 5GHz Wi-Fi bands for boosted low-latency gaming and web speed", "دمج نطاقي 2.4GHz و 5GHz لتسريع تحميل صفحات الإنترنت والألعاب", "Settings > Wi-Fi > Wi-Fi Assistant"],
    [3, "Connection & sharing", "Mi Share", "Cast screen & Private DNS", "High-speed multi-device file transfer, wireless screen casting, and DNS encryption", "مشاركة الملفات بسرعة فائقة وبث الشاشة وتشفير خوادم DNS", "Settings > Connection & sharing"],
    [4, "Wallpapers & Personalization", "Super Wallpapers", "Always-on Display styles", "Interactive 3D animated Super Wallpapers and customizable Always-On lock screens", "خلفيات ثلاثية الأبعاد تفاعلية وتخصيص شاشات القفل الحديثة", "Settings > Wallpapers & Personalization"],
    [5, "Display & Brightness", "Color scheme", "Pro Color / Refresh rate 120Hz", "Professional color calibration engine and adaptive high refresh rate switching", "معايرة دقة الألوان الاحترافية وضبط معدل التحديث السلس", "Settings > Display > Refresh rate"],
    [6, "Sound & vibration", "Dolby Atmos", "Hi-Res Audio & Haptic feedback", "Immersive spatial 3D surround sound and customizable tactical haptic vibrations", "تفعيل الصوت المحيطي المجسم ومعايرة قوة الاهتزازات التكتيكية", "Settings > Sound & vibration > Dolby Atmos"],
    [7, "Notifications & Control center", "Control center style", "HyperOS modern style", "Switch between sleek modern Control Center layout and classic notification shade", "التبديل بين نمط مركز التحكم الحديث وستارة الإشعارات الكلاسيكية", "Settings > Notifications & Control center"],
    [8, "Home screen", "System navigation", "Full screen gestures", "Smooth fluid swipe navigation gestures and home screen app grid configuration", "إيماءات التنقل السلسة وضبط شبكة توزيع أيقونات التطبيقات", "Settings > Home screen > System navigation"],
    [9, "Fingerprints & Face unlock", "Biometrics", "In-display fingerprint & Face data", "Enroll optical/ultrasonic fingerprint scanner and biometric face unlock", "تسجيل بصمات الأصابع البصرية/فوق الصوتية ومسح الوجه لفتح القفل", "Settings > Fingerprints & Face unlock"],
    [10, "Battery & Performance", "Power modes", "Performance / Balanced / Save Battery", "Toggle between maximum hardware performance, balanced mode, and ultra power saver", "التبديل بين وضع الأداء الأقصى ووضع التوفير الفائق وحماية الشحن", "Settings > Battery > Performance"],
    [11, "Special features", "Memory extension", "RAM Plus Virtual RAM (Up to 12GB)", "Expand active RAM memory by allocating up to 12GB from high-speed UFS storage", "زيادة سعة الرام التخيلية مستقطعة من ذاكرة UFS السريعة", "Settings > Additional settings > Memory extension"],
    [12, "Special features", "Game Turbo", "GPU Tuner & Performance optimization", "Optimize GPU rendering pipelines, touch sensitivity, and voice changer for gaming", "رفع كفاءة المعالج الرسومي وتخصيص حساسية اللمس للألعاب الثقيلة", "Settings > Special features > Game Turbo"],
    [13, "Special features", "Floating windows", "Sidebar & Quick ball", "Multitask with resizable floating app windows and quick-access edge sidebar", "فتح التطبيقات في نوافذ عائمة متعددة والوصول السريع عبر الشريط الجانبي", "Settings > Additional settings > Floating windows"],
    [14, "Privacy protection", "Privacy dashboard", "Permissions & Clipboard protection", "Monitor sensor/camera access permissions and protect clipboard data from unauthorized apps", "مراقبة وصول التطبيقات للكاميرا والميكروفون وحماية الحافظة", "Settings > Privacy protection"],
    [15, "AI Services", "AI Subtitles & Eraser", "Generative AI photo edit", "Real-time voice audio transcription and generative AI photo object remover", "ترجمة المقاطع الصوتية الفورية وإزالة العناصر غير المرغوبة من الصور", "Settings > Additional settings > AI Services"],
    [16, "About phone", "HyperOS version", "All specs & Storage details", "View processor clock, RAM, security patch level, and official HyperOS software info", "عرض تفاصيل المعالج، الذاكرة، التحديثات الأمنية وإصدار النظام", "Settings > About phone"],
]

# كتالوج إعدادات أوبو وريلمي ووان بلس (OPPO ColorOS / OxygenOS / Realme UI)
OPPO_COLOROS_SETTINGS = [
    [1, "Mobile network", "5G Settings", "Smart 5G / Data Roaming", "Smart network switching between 4G and 5G to preserve battery consumption", "التبديل التلقائي الذكي بين شبكات 4G و 5G لتوفير استهلاك البطارية", "Settings > Mobile network > 5G Settings"],
    [2, "Wi-Fi", "Wi-Fi Assistant", "Dual-channel network acceleration", "Combine Wi-Fi and mobile data simultaneously for ultra-fast download acceleration", "استخدام الواي فاي وبيانات الهاتف معاً لتحقيق أعلى سرعة تحميل", "Settings > Wi-Fi > Wi-Fi Assistant"],
    [3, "Connection & sharing", "Oppo Share / Quick Share", "Screencast & Link to Windows", "Direct high-speed file sharing and seamless full PC integration with Windows", "مشاركة الملفات السريعة ومزامنة الهاتف بالكامل مع نظام ويندوز", "Settings > Connection & sharing"],
    [4, "Wallpapers & style", "Aquamorphic Design", "AOD Canvas / Edge Lighting", "Aquamorphic visual themes, Always-On Canvas sketches, and notification edge glow", "تخصيص ثيمات النظام المائية وإضاءة الحواف عند وصول الإشعارات", "Settings > Wallpapers & style"],
    [5, "Display & brightness", "Eye comfort & sleep", "Ultra Vision Engine / 120Hz", "AI video color enhancement, eye-comfort blue light filter, and 120Hz fluidity", "تحسين وضوح مقاطع الفيديو وتقليل الضوء الأزرق لراحة العين", "Settings > Display & brightness > Ultra Vision Engine"],
    [6, "Sound & vibration", "O-Haptics", "Spatial Audio / Dolby Atmos", "Realistic 3D spatial haptic feedback engine and immersive cinema sound profiles", "محرك اهتزازات واقعي ثلاثي الأبعاد وصوت محيطي مجسم للأفلام", "Settings > Sound & vibration > O-Haptics"],
    [7, "Security", "Payment protection", "Auto security scan & Anti-fraud", "Secure sandboxed environment for banking apps and proactive anti-fraud scanning", "فحص البيئة الأمنية للتطبيقات المصرفية ومنع هجمات الاحتيال", "Settings > Security > Payment protection"],
    [8, "Privacy", "Private Safe", "App Lock & Hide Apps", "Encrypted private vault to secure confidential photos, documents, and lock sensitive apps", "خزنة مشفرة لحفظ الصور والملفات وتأمين التطبيقات برمز سري مستقل", "Settings > Privacy > Private Safe"],
    [9, "Battery", "Battery Health", "Smart Charging Protection 80%", "Halt charging overnight at 80% to preserve chemical lithium lifespan", "إيقاف الشحن تلقائياً عند 80% أثناء النوم لإطالة العمر الكيميائي للبطارية", "Settings > Battery > Battery Health"],
    [10, "Special features", "Smart Sidebar", "File Dock & Screen Translate", "Floating sidebar for instant screen translation and universal drag-and-drop file dock", "شريط ذكي عائم لترجمة الشاشة الفورية وسحب الملفات بين التطبيقات", "Settings > Special features > Smart Sidebar"],
    [11, "Special features", "Flexible Windows", "Split screen multi-tasking", "Swipe up with 3 fingers to activate split screen and launch multiple apps simultaneously", "تقسيم الشاشة وتشغيل أكثر من تطبيق معاً بسحب 3 أصابع للأعلى", "Settings > Special features > Flexible Windows"],
    [12, "AI Studio & Tools", "AI Eraser 2.0", "AI Best Face & AI Summary", "Fix closed-eye group photos and summarize long web articles using on-device AI", "إصلاح الوجوه المغلقة بالصور وتلخيص المقالات الطويلة بالذكاء الاصطناعي", "Settings > Special features > AI Studio"],
    [13, "RAM Expansion", "Virtual Memory", "Expand RAM (+4GB / +8GB / +12GB)", "Expand available operating RAM by converting up to 12GB of internal storage", "توسيع الذاكرة العشوائية لتسريع تشغيل عشرات التطبيقات بالخلفية", "Settings > About device > RAM > RAM expansion"],
    [14, "About device", "ColorOS / OxygenOS", "Processor, Storage & Official Specs", "Detailed hardware specifications, official OS version, and software update status", "معلومات العتاد وإصدار واجهة المستخدم والتحديثات الرسمية", "Settings > About device"],
]

# كتالوج إعدادات فيفو وآيكو وجوجل بكسل (Vivo OriginOS / Google Pixel / Android Generic)
GENERIC_ANDROID_SETTINGS = [
    [1, "Network & internet", "Internet & SIMs", "5G Auto / eSIM / Hotspot", "Manage mobile data plans, dual eSIM lines, and Wi-Fi hotspot sharing", "إدارة خطوط الاتصال وشرائح eSIM وتوزيع نقطة الاتصال اللاسلكية", "Settings > Network & internet"],
    [2, "Connected devices", "Quick Share & Cast", "Fast Pair & Android Auto", "Instant Bluetooth earphone pairing, smart screen mirroring, and car infotainment integration", "الاقتران الفوري بالسماعات وبث الشاشة للسيارات والتلفزيونات الذكية", "Settings > Connected devices"],
    [3, "Display & Refresh rate", "Smooth Display", "Adaptive 120Hz / Screen Resolution", "Enable adaptive dynamic high refresh rate and adjust screen resolution", "تفعيل معدل التحديث السلس التكيفي وضبط دقة الشاشة", "Settings > Display > Smooth Display"],
    [4, "Wallpaper & style", "Material You", "Dynamic Color Palette / Themed Icons", "Automatically tint system menus, icons, and buttons based on wallpaper color palette", "تلوين أيقونات وقوائم النظام تلقائياً حسب ألوان خلفية الشاشة", "Settings > Wallpaper & style"],
    [5, "Sound & vibration", "Spatial Audio", "Clear Calling & Live Caption", "Filter ambient background noise during calls and generate real-time captions for any audio", "عزل الضوضاء أثناء المكالمات وتوليد ترجمة فورية لأي صوت بالنظام", "Settings > Sound & vibration"],
    [6, "Battery", "Battery Saver", "Adaptive Battery & Reverse Charging", "Smart AI battery life management and reverse wireless charging for accessories", "تحسين استهلاك الطاقة بذكاء والشحن اللاسلكي العكسي للأجهزة الأخرى", "Settings > Battery > Adaptive Battery"],
    [7, "Security & privacy", "Security Hub", "Fingerprint / Face Unlock / Private Space", "Centralized security dashboard and isolated encrypted Private Space for sensitive data", "لوحة أمنية مركزية ومساحة خاصة مشفرة لحماية البيانات الحساسة", "Settings > Security & privacy > Private Space"],
    [8, "Google & AI Assistant", "Gemini & AI Tools", "Circle to Search & Call Assist", "Gemini AI assistant and instant Google search by circling any screen content", "المساعد الذكي Gemini والبحث الفوري برسم دائرة على أي عنصر بالشاشة", "Settings > Google > Gemini"],
    [9, "System & Gestures", "Navigation gestures", "Quick Tap on back / System Updates", "Double-tap back of phone to trigger custom shortcuts and check system updates", "النقر المزدوج على ظهر الهاتف لتنفيذ مهام مخصصة وتحديثات النظام", "Settings > System > Gestures"],
    [10, "About phone", "Device details", "Android Version & Build info", "View hardware model number, serial number, IMEI, and Android build info", "معلومات الطراز والرقم التسلسلي وإصدار نظام التشغيل", "Settings > About phone"],
]


def get_phone_settings_catalog(brand: str, model: str, phone_name: str) -> tuple:
    """
    تحديد كتالوج الإعدادات المناسب وقائمة الأعمدة والبيانات بناءً على ماركة وموديل الهاتف
    مع توفير الوصف باللغتين الإنجليزية والعربية
    """
    full_ident = f"{brand} {model} {phone_name}".lower()

    if "apple" in full_ident or "iphone" in full_ident:
        sheet_title = "iPhone Settings"
        raw_data = APPLE_IOS_SETTINGS
    elif "xiaomi" in full_ident or "poco" in full_ident or "redmi" in full_ident or "hyperos" in full_ident:
        sheet_title = "Xiaomi Settings"
        raw_data = XIAOMI_HYPEROS_SETTINGS
    elif "oppo" in full_ident or "realme" in full_ident or "oneplus" in full_ident or "coloros" in full_ident or "oxygenos" in full_ident:
        sheet_title = "ColorOS Settings"
        raw_data = OPPO_COLOROS_SETTINGS
    elif "samsung" in full_ident or "galaxy" in full_ident or "one ui" in full_ident:
        sheet_title = "Samsung Settings"
        # استخدام كتالوج سامسونج الموسع المعتمد ثنائي اللغة
        raw_data = [
            [1, "Connections", "Wi-Fi", "Intelligent Wi-Fi", "Automatically switch to mobile data when Wi-Fi is unstable and optimize energy", "التحويل لبيانات الهاتف تلقائياً عند ضعف الشبكة وتوفير الطاقة", "Settings > Connections > Wi-Fi > Intelligent Wi-Fi"],
            [1, "Connections", "Wi-Fi", "Advanced settings", "Manage saved Wi-Fi networks and view app Wi-Fi control history logs", "إدارة الشبكات المحفوظة وعرض سجل تحكم التطبيقات بالواي فاي", "Settings > Connections > Wi-Fi > Advanced settings"],
            [1, "Connections", "Wi-Fi Calling", "Calling preference", "Make and receive high-definition voice calls over Wi-Fi when cellular coverage is weak", "إجراء واستقبال المكالمات بوضوح عبر شبكة Wi-Fi عند ضعف التغطية", "Settings > Connections > Wi-Fi Calling"],
            [1, "Connections", "Bluetooth", "Music Share", "Share music playback with friends through paired Bluetooth audio speakers", "مشاركة تشغيل الصوتيات مع الأصدقاء على سماعة البلوتوث المقترنة", "Settings > Connections > Bluetooth > Advanced > Music Share"],
            [1, "Connections", "NFC and contactless payments", "Payment default", "Enable contactless NFC payments and select default digital wallet service", "تفعيل الدفع اللاتلامسي واختيار محفظة الدفع الافتراضية", "Settings > Connections > NFC and contactless payments"],
            [1, "Connections", "SIM manager", "Preferred SIMs / Add eSIM", "Manage physical SIM cards, set primary data lines, and activate digital eSIMs", "إدارة الشرائح وتعيين الشريحة الافتراضية وإضافة خطوط eSIM", "Settings > Connections > SIM manager"],
            [1, "Connections", "Mobile networks", "Network mode (5G/LTE)", "Set preferred network mode, configure data roaming, and manage APN profiles", "تحديد نمط الشبكة وضبط التجوال ونقاط الوصول APN", "Settings > Connections > Mobile networks"],
            [1, "Connections", "Data usage", "Data saver", "Monitor mobile internet consumption and enable Data Saver background restriction", "مراقبة استهلاك الإنترنت وتفعيل نمط توفير البيانات", "Settings > Connections > Data usage > Data saver"],
            [1, "Connections", "Mobile Hotspot and Tethering", "Mobile Hotspot", "Share cellular internet connection as a secure password-protected Wi-Fi network", "مشاركة اتصال الإنترنت كشبكة واي فاي محمية بكلمة مرور", "Settings > Connections > Mobile Hotspot and Tethering > Mobile Hotspot"],
            [1, "Connections", "More connection settings", "Private DNS / VPN", "Encrypt domain name lookups and configure virtual private networks", "تشفير اتصالات الدومين وتفعيل الشبكات الافتراضية الخاصة", "Settings > Connections > More connection settings"],
            [2, "Connected devices", "Quick Share", "Who can share with you", "Send and receive high-res photos and files ultra-fast with Android and PC devices", "إرسال واستقبال الملفات والصور بسرعة فائقة مع أجهزة الأندرويد والكمبيوتر", "Settings > Connected devices > Quick Share"],
            [2, "Connected devices", "Auto switch Buds", "Device switching", "Seamlessly switch Galaxy Buds audio between Samsung phones, tablets, and watches", "التبديل التلقائي لسماعات Galaxy Buds بين أجهزة سامسونج الخاصة بك", "Settings > Connected devices > Auto switch Buds"],
            [2, "Connected devices", "Call & text on other devices", "Device authorization", "Answer phone calls and reply to SMS text messages from your connected Galaxy tablets", "استقبال المكالمات والرسائل على أجهزة التابلت والساعات التابعة لحسابك", "Settings > Connected devices > Call & text on other devices"],
            [2, "Connected devices", "Link to Windows", "Cross-device copy/paste", "Sync notifications, photos, and run mobile phone apps directly on Windows PC", "مزامنة الإشعارات والصور وتشغيل تطبيقات الهاتف على الكمبيوتر", "Settings > Connected devices > Link to Windows"],
            [2, "Connected devices", "Samsung DeX", "Wireless / HDMI DeX", "Transform your phone into a full desktop PC experience on external monitors", "تشغيل واجهة كمبيوتر مكتبية كاملة عند التوصيل بشاشة خارجية", "Settings > Connected devices > Samsung DeX"],
            [2, "Connected devices", "Smart View", "Screen mirroring", "Mirror your phone screen wirelessly to smart TVs and wireless displays", "بث وعرض شاشة الهاتف بالكامل على الشاشات الذكية", "Settings > Connected devices > Smart View"],
            [2, "Connected devices", "Android Auto", "Wireless projection", "Connect phone to car dashboard for GPS navigation, music, and voice assistance", "ربط الهاتف بشاشة السيارة لتشغيل الخرائط والموسيقى والمساعد الصوتي", "Settings > Connected devices > Android Auto"],
            [3, "Galaxy AI", "Phone", "Live translate", "Real-time two-way voice and text translation during phone calls without internet", "ترجمة صوتية ونصية فورية ثنائية الاتجاه للمكالمات بدون إنترنت", "Settings > Galaxy AI > Phone > Live translate"],
            [3, "Galaxy AI", "Samsung Keyboard", "Chat translation & Writing style", "Translate chats and rewrite messages in multiple tones with grammar correction", "ترجمة الدردشات وإعادة صياغة النصوص بنبرات متعددة وتصحيح القواعد", "Settings > Galaxy AI > Samsung Keyboard"],
            [3, "Galaxy AI", "Interpreter", "Language packs", "Real-time face-to-face conversation translation with dual split-screen mode", "ترجمة صوتية فورية للمحادثات المباشرة وجه لوجه بنمط الشاشة المزدوجة", "Settings > Galaxy AI > Interpreter"],
            [3, "Galaxy AI", "Samsung Notes", "Note Assist", "Auto-format, summarize, spell-check, and translate handwritten/typed notes", "تلخيص الملاحظات وتنسيقها وترجمتها آلياً", "Settings > Galaxy AI > Samsung Notes"],
            [3, "Galaxy AI", "Voice Recorder", "Transcript Assist", "Convert voice recordings into speaker-labeled transcripts and smart summaries", "تحويل التسجيل الصوتي لنص مكتوب مع تمييز المتحدثين وتلخيصه", "Settings > Galaxy AI > Voice Recorder"],
            [3, "Galaxy AI", "Samsung Internet", "Browsing Assist", "Summarize and translate lengthy web articles with a single tap in browser", "تلخيص مقالات الويب وترجمتها بضغطة زر داخل المتصفح", "Settings > Galaxy AI > Samsung Internet"],
            [3, "Galaxy AI", "Photo Editor", "Generative Edit", "Relocate, resize, or erase unwanted objects and generate background fills", "تحريك العناصر أو حذفها وإعادة ضبط زوايا الصور وتوليد الخلفيات", "Settings > Galaxy AI > Photo Editor"],
            [3, "Galaxy AI", "Drawing Assist", "Sketch to Image", "Transform simple finger sketches and doodles into finished artistic illustrations", "تحويل الاسكتشات والرسومات البسيطة للوحات وتصميمات فنية متكاملة", "Settings > Galaxy AI > Drawing Assist"],
            [3, "Galaxy AI", "Privacy Control", "Process data only on device", "Enforce on-device AI data processing exclusively to maximize privacy", "قصر معالجة بيانات الذكاء الاصطناعي على المعالج الداخلي لحماية الخصوصية", "Settings > Galaxy AI > Process data only on device"],
            [4, "Modes and Routines", "Modes", "Sleep / Driving / Exercise", "Automate Do Not Disturb, display color modes, and apps based on user activity", "تفعيل بروفايلات تضبط عدم الإزعاج والألوان والتطبيقات حسب نشاطك", "Settings > Modes and Routines > Modes"],
            [4, "Modes and Routines", "Routines", "If / Then automation", "Create powerful conditional automation rules based on time, location, or triggers", "إنشاء سيناريوهات أتمتة تنفذ أوامر محددة عند تحقق شروط معينة", "Settings > Modes and Routines > Routines > +"],
            [5, "Sounds and vibration", "Sound mode", "Sound / Vibrate / Mute", "Switch between ringtone sound, vibration-only, and temporary mute modes", "التبديل بين أنماط الصوت والاهتزاز والكتم المؤقت", "Settings > Sounds and vibration > Sound mode"],
            [5, "Sounds and vibration", "Sound quality and effects", "Dolby Atmos", "Activate 3D spatial surround sound and fine-tune multi-band audio equalizer", "تفعيل الصوت المحيطي المجسم وضبط معادل الصوت Equalizer", "Settings > Sounds and vibration > Sound quality and effects > Dolby Atmos"],
            [5, "Sounds and vibration", "Sound quality and effects", "Adapt Sound", "Calibrate custom audio frequency curves tailored to personal hearing profile", "معايرة ترددات الصوت وتخصيصها لتناسب دقة حاسة السمع الشخصية", "Settings > Sounds and vibration > Sound quality and effects > Adapt Sound"],
            [5, "Sounds and vibration", "Separate app sound", "Audio routing", "Play selected app audio on Bluetooth speaker while routing system audio to phone", "تشغيل صوت تطبيق معين على سماعة بلوتوث وباقي النظام على مكبر الهاتف", "Settings > Sounds and vibration > Separate app sound"],
            [6, "Notifications", "Notification pop-up style", "Brief pop-up settings", "Customize popup notification style and enable dynamic Edge lighting effects", "تخصيص مظهر الإشعار المختصر وتفعيل إضاءة الحواف Edge lighting", "Settings > Notifications > Notification pop-up style > Brief pop-up settings"],
            [6, "Notifications", "Notification summaries", "AI summary", "Summarize incoming notification batches and group chats with on-device AI", "تلخيص الإشعارات المكثفة ومحادثات المجموعات بالذكاء الاصطناعي", "Settings > Notifications > Notification summaries"],
            [6, "Notifications", "Advanced settings", "Notification history", "View and recover past dismissed notifications from the last 24 hours", "استرجاع وقراءة الإشعارات التي تم مسحها بالخطأ خلال 24 ساعة", "Settings > Notifications > Advanced settings > Notification history"],
            [7, "Display", "Motion smoothness", "Adaptive (120Hz)", "Enable dynamic 1Hz-120Hz adaptive refresh rate for ultra-smooth scrolling", "تفعيل التردد التكيفي حتى 120Hz للحصول على سلاسة فائقة", "Settings > Display > Motion smoothness > Adaptive"],
            [7, "Display", "Screen resolution", "QHD+ / FHD+", "Toggle between maximum QHD+ screen clarity and battery-saving FHD+ resolution", "التبديل بين أعلى دقة عرض فائقة QHD+ والدقة القياسية الموفرة FHD+", "Settings > Display > Screen resolution"],
            [7, "Display", "Navigation bar", "Circle to Search", "Instant visual search by long-pressing home button and circling any screen item", "البحث الفوري عن أي عنصر بالضغط المطول ورسم دائرة على الشاشة", "Settings > Display > Navigation bar > Circle to Search"],
            [7, "Display", "Edge panels", "Panels configuration", "Slide-out edge panel for quick access to favorite apps, clipboard, and tools", "شريط جانبي للوصول السريع للتطبيقات المفضلة والحافظة والأدوات", "Settings > Display > Edge panels"],
            [8, "Battery", "Battery protection", "Basic / Adaptive / Maximum", "Protect lithium battery chemistry by capping maximum charge level at 80%", "حماية خلايا البطارية وإيقاف الشحن عند 80% لإطالة عمرها الافتراضي", "Settings > Battery > Battery protection"],
            [8, "Battery", "Charging settings", "Fast charging", "Enable Super Fast wired charging and Fast wireless charging modes", "تفعيل تقنية الشحن السلكي فائق السرعة والشحن اللاسلكي السريع", "Settings > Battery > Charging settings > Fast charging"],
            [8, "Battery", "Wireless power sharing", "Reverse charging", "Use the back of phone to wirelessly charge smartwatches, earbuds, and phones", "استخدام ظهر الهاتف لشحن الساعات والسماعات لاسلكياً", "Settings > Battery > Wireless power sharing"],
            [9, "Wallpaper and style", "Change wallpapers", "Generative AI", "Generate custom AI wallpapers by blending keywords, themes, and artistic styles", "توليد خلفيات فنية مبتكرة باستخدام محرك الذكاء الاصطناعي", "Settings > Wallpaper and style > Change wallpapers > Generative"],
            [9, "Wallpaper and style", "Color palette", "UI dynamic theme", "Extract harmonized color themes from current wallpaper and apply to system UI", "استخراج ألوان متناسقة من الخلفية وتطبيقها على أزرار وقوائم النظام", "Settings > Wallpaper and style > Color palette"],
            [10, "Themes", "Galaxy Themes", "Themes / Icons / AODs", "Download full system themes, custom app icons, and Always On Display art", "تغيير المظهر الشامل للأيقونات والخلفيات وشاشات القفل من متجر جالاكسي", "Settings > Themes"],
            [11, "Home screen", "Home screen grid", "Grid layout (e.g. 5x5)", "Adjust icon grid layout density and number of apps on home and apps screens", "ضبط شبكة توزيع الأيقونات وعدد التطبيقات المعروضة في الشاشة الرئيسية", "Settings > Home screen > Home screen grid"],
            [11, "Home screen", "Hide apps", "App visibility", "Completely hide sensitive installed applications from home screen and app drawer", "إخفاء التطبيقات الحساسة تماماً من الواجهة وقائمة التطبيقات", "Settings > Home screen > Hide apps on Home and Apps screens"],
            [12, "Lock screen and AOD", "Always On Display", "Display mode / Wallpaper", "Display clock and notifications on locked screen with optional dimmed wallpaper", "عرض الساعة والإشعارات على الشاشة المغلقة مع إمكانية عرض الخلفية بذكاء", "Settings > Lock screen and AOD > Always On Display"],
            [12, "Lock screen and AOD", "Now bar", "Live activities", "Interactive live activity bar on lock screen for timers, music, and sports", "شريط تفاعلي أسفل شاشة القفل لمتابعة الأنشطة المباشرة كالمؤقت والموسيقى", "Settings > Lock screen and AOD > Now bar"],
            [12, "Lock screen and AOD", "Extend Unlock", "Smart Lock", "Keep phone unlocked while in trusted locations or connected to trusted smartwatches", "إبقاء الهاتف مفتوحاً في الأماكن الموثوقة أو عند اتصاله بساعتك الذكية", "Settings > Lock screen and AOD > Extend Unlock"],
            [13, "Security and privacy", "Biometrics", "Fingerprints / Face recognition", "Enroll ultrasonic in-display fingerprints and facial recognition biometrics", "تسجيل بصمات الأصابع تحت الشاشة ومسح الوجه لفتح القفل بأمان", "Settings > Security and privacy > Biometrics"],
            [13, "Security and privacy", "Auto Blocker", "Security restrictions", "Block app installs from unknown sources and block malicious USB commands", "منع تثبيت التطبيقات المجهولة وفحص الأوامر والروابط الخبيثة عبر USB", "Settings > Security and privacy > Auto Blocker"],
            [13, "Security and privacy", "Secure Folder", "Knox encrypted container", "Hardware-encrypted Knox private vault for photos, sensitive documents, and cloned apps", "مساحة مشفرة ومنعزلة لحفظ الصور والملفات وتكرار التطبيقات برمز سري", "Settings > Security and privacy > Secure Folder"],
            [14, "Location", "App permissions", "Location access level", "Manage and audit precise and approximate GPS location access permissions", "تصنيف ومراقبة صلاحيات وصول التطبيقات لنظام الملاحة GPS", "Settings > Location > App permissions"],
            [14, "Location", "Location services", "Wi-Fi / Bluetooth scanning", "Improve positioning accuracy by scanning nearby Wi-Fi and Bluetooth devices", "تحسين دقة تحديد الموقع بمسح الشبكات حتى مع إيقاف الواي فاي", "Settings > Location > Location services"],
            [15, "Safety and emergency", "Medical info", "Health conditions / Blood type", "Store emergency medical information and blood type accessible from lock screen", "كتابة فصيلة الدم والحالة الصحية لتظهر للمسعفين على شاشة القفل", "Settings > Safety and emergency > Medical info"],
            [15, "Safety and emergency", "Emergency SOS", "Countdown & Auto call", "Send SOS alert with live GPS coordinates by pressing power button 5 times", "إرسال استغاثة ومشاركة الموقع والاتصال بالإنقاذ بضغط زر التشغيل 5 مرات", "Settings > Safety and emergency > Emergency SOS"],
            [16, "Accounts and backup", "Smart Switch", "Transfer data", "Transfer all photos, contacts, messages, and settings from old phone seamlessly", "نقل البيانات والتطبيقات والرسائل بالكامل من الهاتف القديم سلكياً/لاسلكياً", "Settings > Accounts and backup > Smart Switch"],
            [16, "Accounts and backup", "Samsung Cloud / Google Drive", "Cloud Backup", "Create automated cloud backups of contacts, messages, call logs, and settings", "أخذ نسخة احتياطية سحابية للأسماء والرسائل والملفات والإعدادات", "Settings > Accounts and backup > Back up data"],
            [17, "Google", "All services", "Autofill with Google", "Auto-fill saved passwords, payment cards, and addresses across apps and browser", "التعبئة التلقائية لكلمات المرور المحفوظة والبطاقات البنكية والعناوين", "Settings > Google > All services > Autofill with Google"],
            [17, "Google", "All services", "Devices & sharing", "Manage Matter smart home accessories, Quick Share, and media casting", "إدارة أجهزة المنزل الذكي Matter والمشاركة القريبة وبث الوسائط", "Settings > Google > All services > Devices & sharing"],
            [18, "Advanced features", "Multi window", "Swipe for split screen / Pop-up", "Swipe up with two fingers to split screen or swipe from top corner for pop-up window", "سحب الشاشة بإصبعين لتشغيل تطبيقين معاً أو تحويل التطبيق لنافذة عائمة", "Settings > Advanced features > Multi window"],
            [18, "Advanced features", "Side button", "Double press / Press and hold", "Customize power button double-press and long-press shortcuts for camera or apps", "تخصيص وظيفة الضغط المزدوج والمطول لزر التشغيل لفتح الكاميرا أو التطبيقات", "Settings > Advanced features > Side button"],
            [18, "Advanced features", "Motions and gestures", "Lift to wake / Palm swipe", "Wake screen on lift, double-tap to turn off, and swipe palm to capture screenshot", "إضاءة الشاشة عند الرفع وتمرير راحة اليد لالتقاط صورة الشاشة", "Settings > Advanced features > Motions and gestures"],
            [18, "Advanced features", "Dual Messenger", "App clone", "Run two independent accounts for messaging apps (e.g. WhatsApp) on dual SIMs", "تشغيل نسختين مستقلتين من تطبيقات الدردشة (مثل WhatsApp) برقمين", "Settings > Advanced features > Dual Messenger"],
            [19, "Digital Wellbeing and parental controls", "Screen time & App timers", "Daily usage limits", "Monitor daily screen on time and set app usage limits to maintain focus", "تعيين حدود زمنية يومية لاستخدام التطبيقات ومراقبة وقت الشاشة", "Settings > Digital Wellbeing > App timers"],
            [19, "Digital Wellbeing and parental controls", "Parental controls", "Google Family Link", "Set parental screen time limits, filter web content, and monitor children's app usage", "الإشراف العائلي وتحديد التطبيقات والمحتوى المسموح للأطفال", "Settings > Digital Wellbeing > Parental controls"],
            [20, "Device care", "Performance", "Optimize now", "Clear memory cache, close background apps, and optimize battery with one tap", "تنظيف الذاكرة المؤقتة وإغلاق تطبيقات الخلفية لتحسين السرعة بضغطة زر", "Settings > Device care > Optimize now"],
            [20, "Device care", "Memory", "RAM Plus", "Allocate up to 8GB of virtual RAM using high-speed internal UFS storage", "إضافة ذاكرة عشوائية افتراضية (تصل إلى 8GB) مستقطعة من السعة التخزينية", "Settings > Device care > Memory > RAM Plus"],
            [20, "Device care", "Maintenance mode", "Personal data lock", "Lock personal photos, accounts, and private data when sending phone for repair", "حجب الصور والبيانات والحسابات الشخصية عند إرسال الهاتف للصيانة", "Settings > Device care > Maintenance mode"],
            [21, "Apps", "Choose default apps", "Default Browser / Phone", "Set default applications for web browsing, phone dialer, and digital assistant", "تعيين التطبيقات الافتراضية للتصفح والمكالمات والمساعد الرقمي", "Settings > Apps > Choose default apps"],
            [21, "Apps", "Special access", "Appear on top", "Manage advanced permissions like picture-in-picture and draw over other apps", "إدارة الصلاحيات الخاصة المتقدمة كإمكانية ظهور النوافذ فوق التطبيقات", "Settings > Apps > Special access"],
            [22, "General management", "Language", "App languages", "Configure system display language and assign individual languages per app", "تعيين لغة النظام وتخصيص لغات مستقلة لكل تطبيق على حدة", "Settings > General management > Language"],
            [22, "General management", "Samsung Keyboard settings", "Layout & Predictive text", "Customize keyboard layout, predictive text suggestions, and clipboard history", "تخصيص لغات وتصميم لوحة المفاتيح وتفعيل التنبؤ الذكي والحافظة", "Settings > General management > Samsung Keyboard settings"],
            [22, "General management", "Reset", "Factory data reset / Network reset", "Reset network settings, accessibility settings, or perform full factory data wipe", "إعادة ضبط إعدادات الشبكة أو استعادة ضبط المصنع الشامل للهاتف", "Settings > General management > Reset"],
            [23, "Accessibility", "Vision enhancements", "High contrast / Magnifier", "High contrast themes, color inversion, screen magnifier, and bold font rendering", "تفعيل التباين العالي وعكس الألوان والعدسة المكبرة لتسهيل الرؤية", "Settings > Accessibility > Vision enhancements"],
            [23, "Accessibility", "Hearing enhancements", "Live Transcribe / Amplify sound", "Amplify ambient room audio and generate real-time live captions for conversations", "تضخيم الصوت المحيط وتحويل الكلام المنطوق لنصوص مكتوبة لحظياً", "Settings > Accessibility > Hearing enhancements"],
            [23, "Accessibility", "Advanced settings", "Flash notification", "Flash camera LED light or pulse screen color as a visual notification alert", "تشغيل فلاش الكاميرا أو وميض الشاشة كتنبيه مرئي عند وصول الإشعارات", "Settings > Accessibility > Advanced settings > Flash notification"],
            [24, "Software update", "Download and install", "One UI / OS Updates", "Check, download, and install latest One UI updates and security patches", "التحقق اليدوي من التحديثات الأمنية وإصدارات النظام وتثبيتها", "Settings > Software update > Download and install"],
            [24, "Software update", "Auto download over Wi-Fi", "Wi-Fi update download", "Automatically download system software updates over Wi-Fi to preserve mobile data", "تنزيل ملفات التحديثات تلقائياً عند الاتصال بالواي فاي لتوفير الباقة", "Settings > Software update > Auto download over Wi-Fi"],
            [25, "Tips and user guide", "Welcome to your Galaxy", "AI feature exploration", "Browse interactive guides showcasing Galaxy AI features and One UI hidden tricks", "استعراض بطاقات تفاعلية تشرح أحدث مزايا Galaxy AI وحيل النظام", "Settings > Tips and user guide"],
            [25, "Tips and user guide", "User manual", "Online digital manual", "Open full comprehensive digital user manual in web browser", "فتح دليل المستخدم الرقمي الشامل للهاتف عبر المتصفح", "Settings > Tips and user guide > User manual"],
            [26, "About phone", "Software information", "Build number / Android version", "View Android version, One UI build, and tap Build Number 7 times for Developer Options", "تفاصيل النظام وتفعيل خيارات المطور بالضغط 7 مرات على Build number", "Settings > About phone > Software information"],
            [26, "About phone", "Status information", "IMEI / Serial number / IP", "View hardware serial numbers, dual IMEI codes, and network IP connectivity", "عرض أرقام الهوية العتادية للهاتف وحالة الاتصال والشبكة", "Settings > About phone > Status information"]
        ]
    else:
        sheet_title = "Settings Guide"
        raw_data = GENERIC_ANDROID_SETTINGS

    return sheet_title, raw_data


def export_phone_settings_to_excel(phone_name: str, brand: str = "", model: str = "") -> bytes:
    """
    إنشاء وتصدير شيت إكسيل مخصص ودقيق يشرح جميع قوائم وإعدادات أي هاتف ذكي يتم اختياره
    مع دعم الأعمدة الفرعية (Main Menu, Sub-Menus, Description in English, Description in Arabic, Full Settings Path)
    واتجاه Left-to-Right والتنسيق الملكي الراقي.
    """
    sheet_title, data = get_phone_settings_catalog(brand, model, phone_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet title limit
    # اتجاه الورقة Left-to-Right
    ws.views.sheetView[0].rightToLeft = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # حساب عدد القوائم الفرعية تلقائياً (Dynamic Sub-Menu Columns)
    # الهيكل الأصلي: [Menu #, Main Menu, Sub1, Sub2, ..., Desc_EN, Desc_AR, Path]
    max_len = max(len(row) for row in data) if data else 7
    sub_count = max_len - 5  # طرح (Menu #, Main Menu, Desc_EN, Desc_AR, Path)
    sub_count = max(1, sub_count)

    # الترتيب الجديد: Menu # -> Main Menu -> Full Settings Path -> Sub-Menus -> Description in English -> Description in Arabic
    headers = [
        "Menu #",
        "Main Menu (القائمة الرئيسية)",
        "Full Settings Path (مسار الوصول)"
    ]
    for s_idx in range(1, sub_count + 1):
        headers.append(f"Sub-Menu Level {s_idx} (القائمة الفرعية {s_idx})")
    headers.extend([
        "Description in English (الوصف بالإنجليزية)",
        "Description in Arabic (شرح الوظيفة بالعربية)"
    ])

    # 1. عنوان التقرير الرئيسي (الصف الأول: الماركة والموديل بخط كبير وخلفية أنيقة)
    clean_brand = brand.split(" (")[0].replace("Apple iPhone", "Apple").replace("Xiaomi & Poco & Redmi", "Xiaomi").replace("Vivo & iQOO", "Vivo").strip()
    clean_model = model.strip() if model else phone_name
    if clean_brand and clean_brand.lower() not in clean_model.lower():
        display_title = f"📱 {clean_brand} {clean_model} — Settings & Features Guide (دليل الإعدادات والميزات)"
    else:
        display_title = f"📱 {clean_model} — Settings & Features Guide (دليل الإعدادات والميزات)"

    num_cols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell_title = ws.cell(row=1, column=1, value=display_title)
    cell_title.font = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    cell_title.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 2. الصف الفرعي التوضيحي
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    cell_sub = ws.cell(row=2, column=1, value="📋 Official System Architecture, Settings Paths & Bilingual Feature Descriptions (مسارات الضبط والشرح التفصيلي بالعربية والإنجليزية)")
    cell_sub.font = Font(name="Segoe UI", size=10, italic=True, color="93C5FD")
    cell_sub.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    cell_sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # 3. صف فاصل فارغ
    ws.row_dimensions[3].height = 10

    # 4. تنسيق العناوين والصفوف (الصف الرابع)
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    row_fill_even = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    regular_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    header_row_idx = 4
    ws.row_dimensions[header_row_idx].height = 30
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 5. صفوف البيانات ابتداءً من الصف الخامس
    for row_idx, row_data in enumerate(data, start=5):
        menu_num = row_data[0]
        main_menu = row_data[1]
        desc_en = row_data[-3]
        desc_ar = row_data[-2]
        path_val = row_data[-1]
        subs_existing = list(row_data[2:-3])
        while len(subs_existing) < sub_count:
            subs_existing.append("-")

        # إعادة ترتيب الصف: Menu # -> Main Menu -> Full Settings Path -> Sub-Menus -> Description in English -> Description in Arabic
        reordered_row = [menu_num, main_menu, path_val] + subs_existing + [desc_en, desc_ar]

        ws.row_dimensions[row_idx].height = 26
        is_even = (row_idx % 2 == 0)

        for col_idx, val in enumerate(reordered_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_font if col_idx in [1, 2] else regular_font
            cell.border = thin_border
            if is_even:
                cell.fill = row_fill_even

            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == num_cols:  # Description in Arabic
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ضبط أبعاد الأعمدة تلقائياً
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 45  # Full Settings Path
    for s_idx in range(1, sub_count + 1):
        col_letter = get_column_letter(3 + s_idx)
        ws.column_dimensions[col_letter].width = 28

    desc_en_letter = get_column_letter(len(headers) - 1)
    desc_ar_letter = get_column_letter(len(headers))
    ws.column_dimensions[desc_en_letter].width = 48  # Description in English
    ws.column_dimensions[desc_ar_letter].width = 52  # Description in Arabic

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_s26_settings_to_excel() -> bytes:
    """تصدير شيت إعدادات سامسونج S26 Plus"""
    return export_phone_settings_to_excel("Samsung Galaxy S26 Plus", "Samsung", "Galaxy S26 Plus")


def get_phone_settings_dataframe(phone_name: str, brand: str = "", model: str = "") -> pd.DataFrame:
    """
    إنشاء واسترجاع DataFrame منسق لجميع قوائم وإعدادات الهاتف للعرض المباشر والتفاعلي على الشاشة.
    """
    sheet_title, data = get_phone_settings_catalog(brand, model, phone_name)

    max_len = max(len(row) for row in data) if data else 7
    sub_count = max_len - 5
    sub_count = max(1, sub_count)

    headers = [
        "Menu #",
        "Main Menu (القائمة الرئيسية)",
        "Full Settings Path (مسار الوصول)"
    ]
    for s_idx in range(1, sub_count + 1):
        headers.append(f"Sub-Menu Level {s_idx} (القائمة الفرعية {s_idx})")
    headers.extend([
        "Description in English (الوصف بالإنجليزية)",
        "Description in Arabic (شرح الوظيفة بالعربية)"
    ])

    rows = []
    for row_data in data:
        menu_num = row_data[0]
        main_menu = row_data[1]
        desc_en = row_data[-3]
        desc_ar = row_data[-2]
        path_val = row_data[-1]
        subs_existing = list(row_data[2:-3])
        while len(subs_existing) < sub_count:
            subs_existing.append("-")

        reordered_row = [menu_num, main_menu, path_val] + subs_existing + [desc_en, desc_ar]
        rows.append(reordered_row)

    return pd.DataFrame(rows, columns=headers)


